"""Part 2 Step 4c: export the classification result table as an XLSX file.

Columns (exactly as required by the assignment):

* ``repository_id``
* ``project_type``
* ``project_title``
* ``primary_class``
* ``secondary_class``   (if any)
* ``no_project_files``  (number of files in the project in total)
"""

import os
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_DB = os.path.join(ROOT, "23080363-sq26-classification.db")
DEFAULT_OUT = os.path.join(ROOT, "23080363-sq26-classification.xlsx")

COLUMNS = [
    "repository_id",
    "project_type",
    "project_title",
    "primary_class",
    "secondary_class",
    "no_project_files",
]


def fetch_rows(conn):
    return conn.execute(
        """
        SELECT p.repository_id            AS repository_id,
               pc.project_type            AS project_type,
               p.title                    AS project_title,
               pc.primary_class           AS primary_class,
               pc.secondary_class         AS secondary_class,
               pc.no_project_files        AS no_project_files
        FROM PROJECT_CLASSES pc
        JOIN PROJECTS p ON p.id = pc.project_id
        ORDER BY p.repository_id, pc.project_type, p.id
        """
    ).fetchall()


def export(db_path=DEFAULT_DB, out_path=DEFAULT_OUT):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = fetch_rows(conn)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "classification"

    header_fill = PatternFill("solid", fgColor="4B7BEC")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, name in enumerate(COLUMNS, start=1):
            value = row[name]
            ws.cell(row=r_idx, column=c_idx, value=value if value is not None else "")

    # reasonable column widths
    widths = [14, 16, 60, 42, 42, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"Wrote {len(rows)} rows to {out_path}")
    return out_path


if __name__ == "__main__":
    export()
